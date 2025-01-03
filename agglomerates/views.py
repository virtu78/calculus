
from django.shortcuts import render
import datetime
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import Date, Entry
from .forms import  DateForm, EntryForm
from openpyxl import load_workbook
import xlwt
from django.http import HttpResponse

# Create your views here.
def index(request):
    """Домашняя странтца приложения agglomerate"""
    return render(request, "agglomerates/index.html")




def new_date(request):
    """Определяем группу по для расчетов"""
   
    if request.method != 'POST':
        #Данные не отправлись, создается пустая форма.
        form= DateForm()
    else:
        #Отправлены данные POST, обработать данные.
        form=DateForm(request.POST) 
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('agglomerates:all_dates'))

 

def new_entry(request, id):
    """Определяем добавление нового расчета"""
    date=Date.objects.get(id=id)
    if request.method != 'POST':
        #Данные не отправлись, создается пустая форма.
        form= EntryForm()
    else:
        #Отправлены данные POST, обработать данные.
        form=EntryForm(data=request.POST) 
        if form.is_valid():
            new_entry = form.save(commit=False)
            new_entry.date = date
            new_entry.save()
            return HttpResponseRedirect(reverse('agglomerates:single_date', args=[id]))
        
    context = {'date':date, 'form':form}
    return render(request, 'agglomerates/new_entry.html', context)
    
def all_dates(request):
    """Выводи список дат расчетов"""
    dates=Date.objects.order_by('-date_enter') 
    if request.method != 'POST':
        #Данные не отправлись, создается пустая форма.
        form= DateForm()
    else:
        #Отправлены данные POST, обработать данные.
        form=DateForm(request.POST) 
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('agglomerates:all_dates'))
    
    context = {'form':form, 'dates': dates}
    return render(request, 'agglomerates/all_dates.html', context)

def single_date(request, id):
    """Выводи одну дату расчета и все записи в этот день"""
    date=Date.objects.get(id=id)
    entries=date.entry_set.order_by('-date_added')
    context = {'date':date, 'entries':entries,}
    return render(request, 'agglomerates/date.html', context)

def table(request):
    """Выводи одну дату расчета и все записи в этот день"""
    # entries=Entry.objects.order_by('-date_added') 
    dates=Date.objects.all().prefetch_related('entry_set').order_by('-date_enter')[:30]
    # entries=date.entry_set.order_by('-date_added')   
    context = {'dates':dates}
    return render(request, 'agglomerates/table.html', context)

def edit_actual_output(request, id):
    entry=Entry.objects.get(id=id)
    date=entry.date
    if request.method != 'POST':
        #Данные не отправлись, создается пустая форма.
        form= EntryForm(instance=entry)
    else:
        #Отправлены данные POST, обработать данные.
        form=EntryForm(instance=entry, data=request.POST) 
        if form.is_valid():
            form.save()           
            return HttpResponseRedirect(reverse('agglomerates:table'))

    context = {'entry':entry, 'date':date, 'form':form}
    return render(request, 'agglomerates/edit_actual.html', context)


def export_in_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="agglomerate.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Расчет')

    # style_head = xlwt.easyxf('alignment: wrap True')
    # Sheet header, first row
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True    
    al = xlwt.Alignment()
    al.vert = 0x03      #  Set vertical home
    
    font_style.alignment = al
    font_style.alignment.wrap = 1
    
    columns = [
    'Дата расчета',
    'Содержание титана в агломерате КГОК',
    'Механическая (барабанная) прочность агломерата',
    'Выход в агломерате класса крупности 40+15мм',
    'Расход природного газа',
    'Расход воды в смесительном барабане общий',
    'Расход агломерационной шихты до высева',
    'Температура отходящих газов',
    'Скорость агломашины',
    'Вес постели общий',
    'Содержание титана в концентрате',
    'Содержание в концентрате класса крупности -0,071 мм',
    'Выход в известняке класса крупности +3, мм',
    'Удельный расход топлива',
    'Расход агломерационной шихты после высева',
    'Уровень заполнения бункера постели',
    'Влажность шихты в бункере а/м',
    'Температура в вакуум-камере № 18',
    'Температура в вакуум-камере № 19',
    'Выход аглотсева',
    'Фактическое значение выхода аглоотсева',
    'Абсолютная разница между расчетным и фактическим значениями'
    ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
        # ws.write(row_num, col_num, columns[col_num], style_head)
    # Sheet body, remaining rows
    row_num = 1
    measurement=[' ','%','%','%','м³/т','т/час','т/час','°C','м/мин','т/час','%','%','%','кг/т','т/час','%','%','°C','°C','кг/т','кг/т','Δ']
    for col_num in range(len(measurement)):
        ws.write(row_num, col_num, measurement[col_num], font_style)
        
    font_style = xlwt.XFStyle()
    al = xlwt.Alignment()
    # al.horz = 0x02      #  Set the horizontal
    al.vert = 0x01      #  Set vertical home
    font_style.alignment = al
   
    rows = Entry.objects.all()[:30].values_list(
        "date_added",
        'titanium_content_in_agglomerate',
        'mechanical_strength_of_agglomerate',
        'output_in_agglomerate_of_size_class',
        'specific_consumption_of_natural_gas',
        'water_consumption_in_drum_pelletizer',
        'charge_consumption_before_seeding',
        'exhaust_gas_temperature',
        'velocity_am',
        'jig_bed_weight',
        'content_titan_in_source_ore',
        'content_of_size_class_in_source_ore',
        'output_in_limestone_of_size_class',
        'specific_fuel_consumption',
        'consumption_of_agglomeration_charge_after_seeding',
        'filling_level_of_bed_bin',
        'moisture_content_in_hopper_am',
        'temperature_in_vacuum_chamber_no_eighteen',
        'temperature_in_vacuum_chamber_no_nineteen',
        'output_agglomerate',
        'actual_value_of_agglomeration_output',
        'difference_between_estimated_and_actual_values'        
        )
    rows = [[x.strftime("%d.%m.%Y") if isinstance(x, datetime.datetime) else x for x in row] for row in rows ]
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            ws.col(0).width = 5000
   
    wb.save(response)
    return response


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        skip_first_row = True
        for row in ws.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
            date_enter, tit_con_agg, mech_strength_agg,output_agg_size_class,spec_cons_nat_gas,water_cons_drum,charge_cons_before_seeding,exhaust_gas_tem,vel_am,jig_bed_weight,tit_source_ore,size_class_source_ore,limestone_size_class,spec_fuel_cons,cons_agg_charge_after_seeding,level_of_bed_bin,moisture_hopper_am,temp_vacuum_chamb_eighteen,temp_vacuum_cham_nineteen,output_agglomerate,actual_value,difference_estimated_actual = row
        
            date=Date.objects.create(date_enter=date_enter)
            date.entry_set.create( 
                date_added=datetime.datetime.now().date(),
                titanium_content_in_agglomerate=tit_con_agg,
                mechanical_strength_of_agglomerate=mech_strength_agg,
                output_in_agglomerate_of_size_class=output_agg_size_class,
                specific_consumption_of_natural_gas=spec_cons_nat_gas,   
                water_consumption_in_drum_pelletizer=water_cons_drum,    
                charge_consumption_before_seeding=charge_cons_before_seeding,
                exhaust_gas_temperature=exhaust_gas_tem,
                velocity_am=vel_am,
                jig_bed_weight=jig_bed_weight,
                content_titan_in_source_ore=tit_source_ore,
                content_of_size_class_in_source_ore=size_class_source_ore,
                output_in_limestone_of_size_class=limestone_size_class,
                specific_fuel_consumption=spec_fuel_cons,
                consumption_of_agglomeration_charge_after_seeding=cons_agg_charge_after_seeding,
                filling_level_of_bed_bin=level_of_bed_bin,
                moisture_content_in_hopper_am=moisture_hopper_am,
                temperature_in_vacuum_chamber_no_eighteen=temp_vacuum_chamb_eighteen,
                temperature_in_vacuum_chamber_no_nineteen=temp_vacuum_cham_nineteen,
                output_agglomerate=output_agglomerate,
                actual_value_of_agglomeration_output=actual_value,
                difference_between_estimated_and_actual_values=difference_estimated_actual
                )
        
        return render(request, 'agglomerates/import_success.html')
    return render(request, 'agglomerates/import.html')
    




